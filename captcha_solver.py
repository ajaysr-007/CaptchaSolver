"""
Captcha Solver Script Entry Point.
Runs OCR processing on Excel sheet rows containing captcha image paths or references.
"""
import os
import sys
from captcha_solver import CaptchaSolver

def process_excel(excel_path: str):
    """
    Process image files listed in the first column of the Excel spreadsheet,
    solve the CAPTCHAs, and write output predictions and validation status.
    """
    import openpyxl
    
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    excel_dir = os.path.dirname(os.path.abspath(excel_path))
    
    solver = CaptchaSolver()
    
    first_cell_val = str(sheet.cell(row=1, column=1).value).strip().lower()
    
    start_row = 1
    if first_cell_val in ("path", "image_path", "image", "file_path", "filepath"):
        sheet.cell(row=1, column=3, value="prediction")
        sheet.cell(row=1, column=4, value="match")
        start_row = 2
        
    total_rows = sheet.max_row
    print(f"Starting processing of {total_rows - start_row + 1} rows...")
    
    for r in range(start_row, total_rows + 1):
        img_path_cell = sheet.cell(row=r, column=1).value
        label_cell = sheet.cell(row=r, column=2).value
        
        img_path = str(img_path_cell).strip() if img_path_cell is not None else ""
        
        # Restore if empty, missing or #VALUE!
        if not img_path or img_path == "#VALUE!":
            restored = False
            ref_path = r"C:\Users\Ajay\Desktop\Result_Updated.xlsx"
            if os.path.exists(ref_path):
                try:
                    ref_wb = openpyxl.load_workbook(ref_path)
                    ref_sheet = ref_wb.active
                    ref_val = ref_sheet.cell(row=r, column=1).value
                    if ref_val and str(ref_val).strip() != "#VALUE!":
                        new_img_path = str(ref_val).strip()
                        sheet.cell(row=r, column=1, value=new_img_path)
                        print(f"Row {r}/{total_rows}: Restored path from backup: '{new_img_path}'")
                        img_path = new_img_path
                        restored = True
                except Exception:
                    pass
            
            if not restored:
                if 2 <= r <= 16:
                    new_img_path = f"train/{r - 2}.png"
                elif 17 <= r <= 68:
                    new_img_path = f"train/{r + 32}.png"
                else:
                    new_img_path = ""
                
                if new_img_path:
                    sheet.cell(row=r, column=1, value=new_img_path)
                    print(f"Row {r}/{total_rows}: Inferred path from row pattern: '{new_img_path}'")
                    img_path = new_img_path
                    
        if not img_path:
            sheet.cell(row=r, column=3, value="[Error] Missing image path")
            sheet.cell(row=r, column=4, value="Error")
            print(f"Row {r}/{total_rows}: Missing image path")
            continue
            
        if not os.path.isabs(img_path):
            resolved_path = os.path.join(excel_dir, img_path)
        else:
            resolved_path = img_path
            
        # Try to correct the path if it does not exist
        if not os.path.exists(resolved_path):
            basename = os.path.basename(img_path)
            possible_path = os.path.join(excel_dir, "train", basename)
            if os.path.exists(possible_path):
                new_img_path = f"train/{basename}"
                sheet.cell(row=r, column=1, value=new_img_path)
                print(f"Row {r}/{total_rows}: Corrected path from '{img_path}' to '{new_img_path}'")
                img_path = new_img_path
                resolved_path = possible_path
            else:
                possible_path2 = os.path.join(excel_dir, basename)
                if os.path.exists(possible_path2):
                    new_img_path = basename
                    sheet.cell(row=r, column=1, value=new_img_path)
                    print(f"Row {r}/{total_rows}: Corrected path from '{img_path}' to '{new_img_path}'")
                    img_path = new_img_path
                    resolved_path = possible_path2
            
        print(f"Row {r}/{total_rows}: Processing image '{img_path}' (Resolved: '{resolved_path}')")
        
        if not os.path.exists(resolved_path):
            error_msg = f"[Error] File not found: {resolved_path}"
            sheet.cell(row=r, column=3, value=error_msg)
            sheet.cell(row=r, column=4, value="Error")
            print(f"  -> {error_msg}")
            continue
            
        try:
            result = solver.solve(resolved_path, verbose=False)
            sheet.cell(row=r, column=3, value=result)
            
            # Write comparison in 4th column
            if label_cell is not None:
                is_match = (str(result).strip().upper() == str(label_cell).strip().upper())
                sheet.cell(row=r, column=4, value="Match" if is_match else "Mismatch")
            else:
                sheet.cell(row=r, column=4, value="No Label")
                
            print(f"  -> Solved: {result}")
        except Exception as e:
            error_msg = f"[Error] {type(e).__name__}: {str(e)}"
            sheet.cell(row=r, column=3, value=error_msg)
            sheet.cell(row=r, column=4, value="Error")
            print(f"  -> {error_msg}")
            
    print(f"Saving updated workbook to: {excel_path}")
    wb.save(excel_path)
    print("Done!")

def process_results_excel(excel_path: str):
    """
    Process the images in the train folder in natural sorted order,
    and write the prediction results sequentially into Column D of Results.xlsx.
    """
    import openpyxl
    import re
    
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    excel_dir = os.path.dirname(os.path.abspath(excel_path))
    train_dir = os.path.join(excel_dir, "train")
    
    if not os.path.exists(train_dir):
        print(f"Error: train directory not found at {train_dir}")
        return
        
    def natural_sort_key(s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]
        
    files = [f for f in os.listdir(train_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    files.sort(key=natural_sort_key)
    
    print(f"Found {len(files)} images in train folder.")
    
    sheet.cell(row=1, column=4, value="Generated Output")
    
    solver = CaptchaSolver()
    
    for idx, filename in enumerate(files):
        row = idx + 2
        img_path = os.path.join(train_dir, filename)
        print(f"Row {row}: Processing image '{filename}'")
        
        try:
            result = solver.solve(img_path, verbose=False)
            sheet.cell(row=row, column=4, value=result)
            print(f"  -> Solved: {result}")
        except Exception as e:
            error_msg = f"[Error] {type(e).__name__}: {str(e)}"
            sheet.cell(row=row, column=4, value=error_msg)
            print(f"  -> {error_msg}")
            
    print(f"Saving updated workbook to: {excel_path}")
    wb.save(excel_path)
    print("Done!")

if __name__ == "__main__":
    default_excel = "Results.xlsx"
    if not os.path.exists(default_excel):
        default_excel = "Result.xlsx"
    if not os.path.exists(default_excel):
        default_excel = r"C:\Users\Ajay\Desktop\CaptchaSolver\Results.xlsx"
    if not os.path.exists(default_excel):
        default_excel = r"C:\Users\Ajay\Desktop\CaptchaSolver\Result.xlsx"
    if not os.path.exists(default_excel):
        default_excel = r"C:\Users\Ajay\Desktop\Result.xlsx"
        
    if len(sys.argv) > 1 and sys.argv[1].lower().endswith(".xlsx"):
        excel_path = sys.argv[1]
    elif os.path.exists(default_excel):
        excel_path = default_excel
    else:
        excel_path = ""
        
    if excel_path:
        if excel_path.lower().endswith("results.xlsx"):
            process_results_excel(excel_path)
        else:
            process_excel(excel_path)
    else:
        image_path = sys.argv[1] if len(sys.argv) > 1 else "captcha.png"
        solver = CaptchaSolver()
        result = solver.solve(image_path, verbose=True)
        print("=" * 50)
        print(f"CAPTCHA TEXT: {result}")
        print("=" * 50)

